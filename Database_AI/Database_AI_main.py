from mtranslate import translate
import openpyxl
import time
import re



# Python Database Function to extract data, translate data, sort data, validate data and match ultimately them.
# The function is called using the main function.
# The function automatically handles reading and writing to an excel file
def database_function():a
    parent_list_one = []
    parent_list_two = []
    cust_list_one = []
    cust_list_two = []
    wb = openpyxl.Workbook()  # opening the excel workbook
    sheet = wb.active
    record_name = input("Please Enter the name of the database file along with the extension(.xlsx) : ")
    file_name = input("Please Enter a name for saving the file along with the extension(.xlsx) : ")
    print("Please wait while the program executes....")
    print("Expected execution time : 400 - 500 seconds")
    excel_doc = openpyxl.load_workbook(record_name)
    sheet_obj = excel_doc.active
    max_row = sheet_obj.max_row
    for i in range(1, max_row + 1):  # for loop to get the parent_names and perform translation
        cell_obj = sheet_obj.cell(row=i, column=1)
        i = cell_obj.value
        count = 0
        state = 'false'
        count += 1
        if (i != None and len(i) > 0):
            words = i.split(" ")
            for word in words:
                if not ord(word[0]) < 126 or ord(word[0]) == 34:
                    trans = translate(i)  # Custom translate API that performs the translation
                    trans = trans.upper()
                    parent_list_one.append(trans)
                    parent_list_two.append(i.upper())
                    break
                else:
                    parent_list_one.append(i.upper())
                    parent_list_two.append(i.upper())
                    break
        elif (i == None):
            parent_list_one.append("UNMATCHED")
            parent_list_two.append("UNMATCHED")

    for i in range(1, max_row + 1):   # for loop to get the customer_names and perform translation
        cell_obj = sheet_obj.cell(row=i, column=2)
        i = cell_obj.value
        count = 0
        state = 'false'
        count += 1
        if (i != None):
            words = i.split(" ")
            for word in words:
                if not ord(word[0]) < 126 or ord(word[0]) == 34:
                    trans = translate(i)  # Custom translate API that performs the translation
                    trans = trans.upper()
                    cust_list_one.append(trans)
                    cust_list_two.append(i.upper())
                    break
                else:
                    cust_list_one.append(i.upper())
                    cust_list_two.append(i.upper())
                    break
        elif (i == None):
            cust_list_one.append("UNMACTHED")
            cust_list_two.append("UNMATCHED")

    gb_list_one = []
    for i in range(1, max_row + 1):  # for loop to get the global_names.
        cell_obj = sheet_obj.cell(row=i, column=3)
        i = cell_obj.value
        count = 0
        count += 1
        if (i != None):
            words = i.split(" ")
            for word in words:
                if not ord(word[0]) < 126 or ord(word[0]) == 34:
                    gb_list_one.append(i)
                    break
                else:
                    gb_list_one.append(i)
                    break
        elif (i == None):
            gb_list_one.append("UNMATCHED")

    db_list_one = []
    for i in range(1, max_row + 1):  # for loop to get the database_names.
        cell_obj = sheet_obj.cell(row=i, column=4)
        i = cell_obj.value
        count = 0
        count += 1
        if (i != None):
            words = i.split(" ")
            for word in words:
                if not ord(word[0]) < 126 or ord(word[0]) == 34:
                    db_list_one.append(i.upper())
                    break
                else:
                    db_list_one.append(i.upper())
                    break
        elif (i == None):
            db_list_one.append("UNMATCHED")

    cpid_list_one = []
    for i in range(1, max_row + 1):  # for loop to get the company_ID.
        cell_obj = sheet_obj.cell(row=i, column=5)
        i = cell_obj.value
        count = 0
        count += 1
        if (i != None):
            cpid_list_one.append(i)
        elif (i == None):
            cpid_list_one.append("UNMATCHED")

    dict = {}
    for i in range(0, len(parent_list_one)):  # loop to create dictionary using p_name as key and global_name as value.
        if (parent_list_one[i] not in dict.keys() or dict[parent_list_one[i]] == "UNMATCHED"):
            dict[parent_list_one[i]] = gb_list_one[i]

    dict_two = {}
    for i in range(0, len(parent_list_one)):  # loop to create dictionary using p_name as key and cust_name as value.
        if (parent_list_one[i] not in dict_two.keys() or dict_two[parent_list_one[i]] == "UNMATCHED"):
            dict_two[parent_list_one[i]] = cust_list_one[i]

    dict_three = {}
    for i in range(0, len(parent_list_one)):  # loop to create dictionary using p_name as key and cpid as value.
        if (parent_list_one[i] not in dict_three.keys() or dict_three[parent_list_one[i]] == "UNMATCHED"):
            dict_three[parent_list_one[i]] = cpid_list_one[i]

    dict_four = {}
    for i in range(0, len(parent_list_one)):  # loop to create dictionary using p_name as key and db_name as value.
        if (parent_list_one[i] not in dict_four.keys() or dict_four[parent_list_one[i]] == "UNMATCHED"):
            dict_four[parent_list_one[i]] = db_list_one[i]

    dict_five = {}
    for i in range(0, len(parent_list_one)):  # loop to create dictionary using p_name as key and p_name_2 as value.
        if (parent_list_one[i] not in dict_five.keys() or dict_five[parent_list_one[i]] == "UNMATCHED"):
            dict_five[parent_list_one[i]] = parent_list_two[i]

    dict_six = {}
    for i in range(0, len(parent_list_one)):  # loop to create dictionary using p_name as key and cust_name_2 as value.
        if (parent_list_one[i] not in dict_six.keys() or dict_six[parent_list_one[i]] == "UNMATCHED"):
            dict_six[parent_list_one[i]] = cust_list_two[i]

    matched_records = {}
    unmatched_records = {}
    for i in dict.keys():  # for loop to differentiate matched and unmatched records
        if (dict[i].strip() != 'UNMATCHED'):
            matched_records[i] = dict[i]
        elif (dict[i].strip() == 'UNMATCHED'):
            unmatched_records[i] = dict[i]

    matched_records_two = {}
    unmatched_records_two = {}
    for i in unmatched_records.keys():  # for loop to differentiate matched and unmatched records
        if (dict_two[i].strip() != 'UNMATCHED'):
            matched_records_two[i] = dict_two[i]
        elif (dict_two[i].strip() == 'UNMATCHED'):
            unmatched_records_two[i] = dict_two[i]

    wb_one = openpyxl.Workbook()
    sheet_one = wb_one.active
    row_one = 1
    for i in matched_records.keys():  # for loop to write the matched records using global_name to the excel file

        c_one = sheet_one.cell(row=row_one, column=1)
        c_two = sheet_one.cell(row=row_one, column=2)
        c_three = sheet_one.cell(row=row_one, column=3)
        c_four = sheet_one.cell(row=row_one, column=4)
        c_five = sheet_one.cell(row=row_one, column=5)
        c_six = sheet_one.cell(row=row_one, column=6)
        c_seven = sheet_one.cell(row=row_one, column=7)
        c_eight = sheet_one.cell(row=row_one, column=8)
        c_one.value = dict_five[i]
        x = i
        c_two.value = re.sub("[\(\[].*?[\)\]]", "", x)
        c_three.value = dict_six[i]
        x = dict_two[i]
        c_four.value = re.sub("[\(\[].*?[\)\]]", "", x)
        c_five.value = matched_records[i]
        c_six.value = dict_four[i]
        c_seven.value = dict_three[i]
        if ((c_one.value).strip() == (c_six.value).strip()):
            c_eight.value = "high"
        else:
            c_eight.value = "low"
        row_one += 1
    count_final = row_one

    row_one = count_final + 2

    c_initial = sheet_one.cell(row=count_final + 1, column=1)
    c_initial.value = "*********************************************************************************SORTED BY CLOSEST************************************************************************************************************"
    row_one = count_final + 2
    for i in matched_records.keys():  # for loop to write the matched records using closest_name to the excel file

        c_one = sheet_one.cell(row=row_one, column=1)
        c_two = sheet_one.cell(row=row_one, column=2)
        c_three = sheet_one.cell(row=row_one, column=3)
        c_four = sheet_one.cell(row=row_one, column=4)
        c_five = sheet_one.cell(row=row_one, column=5)
        c_six = sheet_one.cell(row=row_one, column=6)
        c_seven = sheet_one.cell(row=row_one, column=7)
        c_eight = sheet_one.cell(row=row_one, column=8)
        c_one.value = dict_five[i]
        x = i
        c_two.value = re.sub("[\(\[].*?[\)\]]", "", x)
        c_three.value = dict_six[i]
        x = dict_two[i]
        c_four.value = re.sub("[\(\[].*?[\)\]]", "", x)
        c_five.value = matched_records[i]
        c_six.value = dict_four[i]
        c_seven.value = dict_three[i]
        if ((c_one.value).strip() == (c_six.value).strip()):
            c_eight.value = "high"
        else:
            c_eight.value = "low"
        row_one += 1

    # Print details regarding records

    print("Total Valid Matched records : ", (len(matched_records) + len(matched_records_two)))
    print("Total Duplicate records eliminated : ",
          (len(parent_list_one) - (len(matched_records) + len(matched_records_two))))

    wb_one.save(file_name)

    ### Code to output an excel file containing the unmatched records. Use when you have unmatched records.

    # # row_one = 1
    # # col_one = 1
    # # col_two = 2
    # # wb_two = openpyxl.Workbook()
    # # sheet_two = wb_two.active
    # # for i in unmatched_records_two.keys():
    # #     c_one = sheet_two.cell(row=row_one, column=col_one)
    # #     c_two = sheet_two.cell(row=row_one, column=col_two)
    # #     c_one.value = i
    # #     c_two.value = unmatched_records_two[i]
    # #     row_one += 1
    # # wb_two.save("C:\\Users\\ajt29\\Desktop\\C121.xlsx")
    # #
    # # track_list = []
    # # for i in matched_records.keys():
    # #     track_list.append(i)

# Main function to call the database_function()
def main():
    database_function()

# Calls the main function which in turn calls the database_function()
if __name__ == '__main__':
    start_time = time.clock()  # time module to track the time of execution
    c = main()
    print("Time taken for the program to execute records is", round(time.clock() - start_time), "seconds")
